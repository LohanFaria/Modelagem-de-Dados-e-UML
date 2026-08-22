import io
import openpyxl
from datetime import date
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import PermissionDenied
from apps.clientes.models import Cliente
from apps.investimentos.models import Investimento
from apps.relacionamento.models import Contato
from . import queries


def eh_gestor_ou_auditor(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=["Gestor", "Auditor"]).exists()


@login_required
@user_passes_test(eh_gestor_ou_auditor)
def painel_relatorios(request):
    qualidade = queries.painel_qualidade_dados()
    investimentos = queries.total_investido_por_tipo()
    produtividade = queries.produtividade_contatos()
    clientes_inativos = queries.clientes_inativos(dias=30)
    recentes_investimentos = Investimento.objects.select_related("cliente", "tipo").order_by("-data_aplicacao")[:5]
    recentes_contatos = Contato.objects.select_related("cliente", "funcionario", "forma", "assunto").order_by("-data_contato")[:5]

    return render(request, "relatorios/painel.html", {
        "qualidade": qualidade,
        "investimentos": investimentos,
        "produtividade": produtividade,
        "clientes_inativos": clientes_inativos,
        "recentes_investimentos": recentes_investimentos,
        "recentes_contatos": recentes_contatos,
    })




@login_required
@user_passes_test(eh_gestor_ou_auditor)
def carteira_cliente(request):
    cliente_id = request.GET.get("cliente_id")
    dados_carteira = None
    if cliente_id:
        dados_carteira = queries.carteira_completa_cliente(cliente_id)
    
    clientes_lista = Cliente.objects.only("id", "nome", "cpf").order_by("nome")
    return render(request, "relatorios/carteira.html", {
        "dados_carteira": dados_carteira,
        "clientes_lista": clientes_lista,
        "cliente_selecionado": int(cliente_id) if cliente_id and cliente_id.isdigit() else None,
    })


@login_required
@user_passes_test(eh_gestor_ou_auditor)
def investimentos_por_tipo(request):
    dados = queries.total_investido_por_tipo()
    return render(request, "relatorios/investimentos_tipo.html", {"dados": dados})


@login_required
@user_passes_test(eh_gestor_ou_auditor)
def evolucao_saldo(request):
    cliente_id = request.GET.get("cliente_id")
    saldos = queries.evolucao_saldo_contas(cliente_id)
    clientes_lista = Cliente.objects.only("id", "nome", "cpf").order_by("nome")
    return render(request, "relatorios/evolucao_saldo.html", {
        "saldos": saldos,
        "clientes_lista": clientes_lista,
    })


@login_required
@user_passes_test(eh_gestor_ou_auditor)
def produtividade_funcionarios(request):
    dados = queries.produtividade_contatos()
    return render(request, "relatorios/produtividade.html", {"dados": dados})


@login_required
@user_passes_test(eh_gestor_ou_auditor)
def clientes_sem_contato(request):
    dias = int(request.GET.get("dias", 30))
    clientes = queries.clientes_inativos(dias=dias)
    return render(request, "relatorios/reativacao.html", {
        "clientes": clientes,
        "dias": dias,
    })


@login_required
@user_passes_test(eh_gestor_ou_auditor)
def qualidade_dados(request):
    dados = queries.painel_qualidade_dados()
    return render(request, "relatorios/qualidade_dados.html", {"dados": dados})


# ==============================================================================
# EXPORTAÇÃO DE RELATÓRIOS PARA EXCEL (.XLSX)
# ==============================================================================

@login_required
@user_passes_test(eh_gestor_ou_auditor)
def exportar_investimentos_xlsx(request):
    dados = queries.total_investido_por_tipo()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Investimentos por Tipo"

    # Cabeçalho
    ws.append(["Tipo de Investimento", "Volume Alocado (R$)", "Quantidade de Aplicações", "Representatividade (%)"])

    # Dados
    for item in dados["itens"]:
        ws.append([item["tipo"], float(item["total"]), item["qtd"], float(item["percentual"])])

    # Total
    ws.append(["TOTAL GERAL", float(dados["total_geral"]), sum(i["qtd"] for i in dados["itens"]), 100.0])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="relatorio_investimentos_por_tipo.xlsx"'
    return response


@login_required
@user_passes_test(eh_gestor_ou_auditor)
def exportar_reativacao_xlsx(request):
    dias = int(request.GET.get("dias", 30))
    clientes = queries.clientes_inativos(dias=dias)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes para Reativação"

    ws.append(["Nome do Cliente", "CPF", "Último Contato Registrado"])

    for cli in clientes:
        dt_contato = cli.ultimo_contato.strftime("%d/%m/%Y") if cli.ultimo_contato else "Nunca contatado"
        ws.append([cli.nome, cli.cpf, dt_contato])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="clientes_sem_contato_{dias}_dias.xlsx"'
    return response


@login_required
@user_passes_test(eh_gestor_ou_auditor)
def exportar_produtividade_xlsx(request):
    dados = queries.produtividade_contatos()

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Por Funcionário"
    ws1.append(["Funcionário", "Total de Atendimentos"])
    for item in dados["por_funcionario"]:
        ws1.append([item["funcionario__nome"], item["total"]])

    ws2 = wb.create_sheet(title="Por Assunto")
    ws2.append(["Assunto / Produto", "Total de Demandas"])
    for item in dados["por_assunto"]:
        ws2.append([item["assunto__nome"], item["total"]])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="relatorio_produtividade_equipe.xlsx"'
    return response
