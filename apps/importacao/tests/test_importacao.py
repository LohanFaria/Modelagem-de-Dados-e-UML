import pytest
import openpyxl
from decimal import Decimal
from datetime import date
from django.core.management import call_command
from apps.clientes.models import Cliente, ContaBancaria
from apps.investimentos.models import Investimento
from apps.relacionamento.models import Contato, ContatoQuarentena
from apps.importacao.services import ImportadorPlanilhasService
from apps.importacao.parsers import (
    parse_cpf,
    parse_telefones,
    parse_emails,
    parse_moeda,
    parse_data,
)


def test_parsers_unitarios():
    # Moeda
    assert parse_moeda("R$ 15.000,00") == Decimal("15000.00")
    assert parse_moeda("15000.50") == Decimal("15000.50")
    assert parse_moeda("  1.250,50  ") == Decimal("1250.50")
    assert parse_moeda(None) == Decimal("0.00")
    assert parse_moeda("") == Decimal("0.00")

    # Data
    assert parse_data("01/08/2026") == date(2026, 8, 1)
    assert parse_data("2026-08-01") == date(2026, 8, 1)
    assert parse_data(date(2026, 8, 1)) == date(2026, 8, 1)
    assert parse_data("data_invalida") is None
    assert parse_data(None) is None

    # CPF
    assert parse_cpf("529.982.247-25") == "529.982.247-25"
    assert parse_cpf("52998224725") == "529.982.247-25"

    # Telefones múltiplos
    tels = parse_telefones("(11) 91111-1111 / (11) 92222-2222 , 1133334444")
    assert len(tels) == 3
    assert "(11) 91111-1111" in tels


    # E-mails múltiplos
    emails = parse_emails("ana@xpto.com / ana.silva@gmail.com, contato@ana.com")
    assert len(emails) == 3
    assert "ana@xpto.com" in emails


@pytest.mark.django_db
def test_importador_com_planilha_mock(tmp_path):
    arquivo = tmp_path / "teste_xpto.xlsx"
    wb = openpyxl.Workbook()

    # Aba Clientes
    ws_cli = wb.active
    ws_cli.title = "Clientes"
    ws_cli.append(["Nome", "CPF", "Telefone", "Email", "Banco", "Agência", "Conta", "Saldo", "Data"])
    ws_cli.append(["Marcos Vinicius", "529.982.247-25", "(11) 91111-1111 / (11) 92222-2222", "marcos@test.com", "Banco XPTO", "0001", "10001-1", "15000.00", "01/08/2026"])
    ws_cli.append(["Marcos Vinicius", "529.982.247-25", "(11) 91111-1111", "marcos@test.com", "Banco do Brasil", "0002", "20002-2", "25000.00", "01/08/2026"])

    # Aba Investimentos
    ws_inv = wb.create_sheet(title="Investimentos")
    ws_inv.append(["CPF", "Tipo", "Valor", "Data"])
    ws_inv.append(["529.982.247-25", "CDB XPTO 110%", "R$ 50.000,00", "15/01/2026"])
    ws_inv.append(["999.888.777-66", "Ações B3", "R$ 20.000,00", "20/01/2026"])  # CPF inexistente

    # Aba Contatos
    ws_ct = wb.create_sheet(title="Contatos")
    ws_ct.append(["CPF", "Funcionário", "Forma", "Assunto", "Data", "Observação"])
    ws_ct.append(["529.982.247-25", "Atendente 1", "WhatsApp", "Aporte", "05/08/2026", "Contato ok"])
    ws_ct.append(["999.888.777-66", "Atendente 2", "Telefone", "Dúvida", "06/08/2026", "Órfão"])

    wb.save(arquivo)

    # 1. Teste Dry Run
    service = ImportadorPlanilhasService(str(arquivo))
    rel_dry = service.executar(dry_run=True)
    assert rel_dry["clientes"]["clientes_criados"] == 1
    assert Cliente.objects.count() == 0  # Rollback funcionou

    # 2. Teste Execução Real
    rel_real = service.executar(dry_run=False)
    assert rel_real["clientes"]["clientes_criados"] == 1
    assert rel_real["clientes"]["contas_criadas"] == 2
    assert rel_real["investimentos"]["criados"] == 1
    assert rel_real["investimentos"]["ignorados"] == 1
    assert rel_real["contatos"]["quarentenados"] == 1

    cliente = Cliente.objects.get(cpf="529.982.247-25")
    assert cliente.contas.count() == 2
    assert cliente.telefones.count() == 2
    assert cliente.investimentos.count() == 1
    assert ContatoQuarentena.objects.filter(motivo="CLIENTE_INEXISTENTE").count() == 1

    # 3. Idempotência: reexecutar não duplica clientes nem contas
    service.executar(dry_run=False)
    assert Cliente.objects.count() == 1
    assert ContaBancaria.objects.filter(cliente=cliente).count() == 2


@pytest.mark.django_db
def test_management_command_importar_planilhas(tmp_path):
    arquivo = tmp_path / "cmd_xpto.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["Nome", "CPF", "Telefone", "Email", "Banco", "Agência", "Conta", "Saldo", "Data"])
    ws.append(["Camila Torres", "529.982.247-25", "(11) 99999-8888", "camila@test.com", "Nubank", "0001", "77777-1", "5000.00", "01/08/2026"])
    wb.save(arquivo)

    call_command("importar_planilhas", arquivo=str(arquivo), sem_validacao_cpf=True)
    assert Cliente.objects.filter(nome="Camila Torres").exists()
