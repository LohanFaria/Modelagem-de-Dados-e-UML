from decimal import Decimal
from datetime import date
from django.db import transaction
import openpyxl
from apps.clientes.models import Cliente, Telefone, Email, ContaBancaria, SaldoHistorico
from apps.investimentos.models import TipoInvestimento, Investimento
from apps.relacionamento.models import Funcionario, FormaContato, Assunto, Contato, ContatoQuarentena
from .parsers import parse_cpf, parse_telefones, parse_emails, parse_moeda, parse_data


class ImportadorPlanilhasService:
    def __init__(self, arquivo_path: str, sem_validacao_cpf: bool = False):
        self.arquivo_path = arquivo_path
        self.sem_validacao_cpf = sem_validacao_cpf
        self.relatorio = {
            "clientes": {"lidas": 0, "clientes_criados": 0, "contas_criadas": 0, "saldos_criados": 0},
            "investimentos": {"lidas": 0, "criados": 0, "ignorados": 0},
            "contatos": {"lidas": 0, "criados": 0, "quarentenados": 0},
            "quarentena_motivos": {},
        }

    def executar(self, dry_run: bool = False):
        self.relatorio = {
            "clientes": {"lidas": 0, "clientes_criados": 0, "contas_criadas": 0, "saldos_criados": 0},
            "investimentos": {"lidas": 0, "criados": 0, "ignorados": 0},
            "contatos": {"lidas": 0, "criados": 0, "quarentenados": 0},
            "quarentena_motivos": {},
        }
        wb = openpyxl.load_workbook(self.arquivo_path, data_only=True)
        
        try:
            with transaction.atomic():
                if "Clientes" in wb.sheetnames:
                    self._importar_clientes(wb["Clientes"])
                if "Investimentos" in wb.sheetnames:
                    self._importar_investimentos(wb["Investimentos"])
                if "Contatos" in wb.sheetnames:
                    self._importar_contatos(wb["Contatos"])
                
                if dry_run:
                    transaction.set_rollback(True)
        finally:
            wb.close()
            
        return self.relatorio

    def _importar_clientes(self, sheet):
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return
        headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        
        def idx(col_name):
            for i, h in enumerate(headers):
                if col_name in h:
                    return i
            return None

        idx_nome = idx("nome")
        idx_cpf = idx("cpf")
        idx_tel = idx("telefone")
        idx_email = idx("email") or idx("e-mail")
        idx_banco = idx("banco")
        idx_agencia = idx("agência") or idx("agencia")
        idx_conta = idx("conta")
        idx_saldo = idx("saldo")
        idx_data_saldo = idx("data")

        for row in rows[1:]:
            if not any(row):
                continue
            self.relatorio["clientes"]["lidas"] += 1
            
            nome = str(row[idx_nome]).strip() if idx_nome is not None and row[idx_nome] else "Cliente Sem Nome"
            cpf_raw = str(row[idx_cpf]).strip() if idx_cpf is not None and row[idx_cpf] else ""
            cpf = parse_cpf(cpf_raw)
            if not cpf:
                continue

            cliente, criado = Cliente.objects.get_or_create(
                cpf=cpf,
                defaults={"nome": nome}
            )
            if criado:
                self.relatorio["clientes"]["clientes_criados"] += 1

            # Telefones
            if idx_tel is not None and row[idx_tel]:
                telefones = parse_telefones(row[idx_tel])
                for t in telefones:
                    Telefone.objects.get_or_create(cliente=cliente, numero=t)

            # Emails
            if idx_email is not None and row[idx_email]:
                emails = parse_emails(row[idx_email])
                for e in emails:
                    Email.objects.get_or_create(cliente=cliente, endereco=e)

            # Conta Bancária
            banco = str(row[idx_banco]).strip() if idx_banco is not None and row[idx_banco] else None
            agencia = str(row[idx_agencia]).strip() if idx_agencia is not None and row[idx_agencia] else None
            conta_num = str(row[idx_conta]).strip() if idx_conta is not None and row[idx_conta] else None

            if banco and agencia and conta_num:
                conta_obj, c_criada = ContaBancaria.objects.get_or_create(
                    cliente=cliente,
                    banco=banco,
                    agencia=agencia,
                    conta=conta_num,
                )
                if c_criada:
                    self.relatorio["clientes"]["contas_criadas"] += 1

                # Saldo Histórico
                if idx_saldo is not None and row[idx_saldo] is not None:
                    saldo_val = parse_moeda(row[idx_saldo])
                    data_val = parse_data(row[idx_data_saldo]) if idx_data_saldo is not None else date.today()
                    if data_val:
                        saldo_obj, s_criado = SaldoHistorico.objects.get_or_create(
                            conta=conta_obj,
                            data_saldo=data_val,
                            defaults={"saldo": saldo_val}
                        )
                        if s_criado:
                            self.relatorio["clientes"]["saldos_criados"] += 1

    def _importar_investimentos(self, sheet):
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return
        headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

        def idx(col_name):
            for i, h in enumerate(headers):
                if col_name in h:
                    return i
            return None

        idx_cpf = idx("cpf")
        idx_tipo = idx("tipo") or idx("investimento")
        idx_valor = idx("valor")
        idx_data = idx("data")

        for row in rows[1:]:
            if not any(row):
                continue
            self.relatorio["investimentos"]["lidas"] += 1

            cpf = parse_cpf(row[idx_cpf]) if idx_cpf is not None and row[idx_cpf] else ""
            tipo_nome = str(row[idx_tipo]).strip() if idx_tipo is not None and row[idx_tipo] else "Outros"
            valor = parse_moeda(row[idx_valor]) if idx_valor is not None else Decimal("0.00")
            data_app = parse_data(row[idx_data]) if idx_data is not None else None

            cliente = Cliente.objects.filter(cpf=cpf).first()
            if not cliente:
                self.relatorio["investimentos"]["ignorados"] += 1
                continue

            tipo_obj, _ = TipoInvestimento.objects.get_or_create(nome=tipo_nome)
            Investimento.objects.get_or_create(
                cliente=cliente,
                tipo=tipo_obj,
                valor_investido=valor,
                data_aplicacao=data_app,
            )
            self.relatorio["investimentos"]["criados"] += 1

    def _importar_contatos(self, sheet):
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return
        headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

        def idx(col_name):
            for i, h in enumerate(headers):
                if col_name in h:
                    return i
            return None

        idx_cpf = idx("cpf")
        idx_func = idx("funcionário") or idx("funcionario")
        idx_forma = idx("forma")
        idx_assunto = idx("assunto") or idx("produto")
        idx_data = idx("data")
        idx_obs = idx("observação") or idx("observacao")

        for row in rows[1:]:
            if not any(row):
                continue
            self.relatorio["contatos"]["lidas"] += 1

            cpf = parse_cpf(row[idx_cpf]) if idx_cpf is not None and row[idx_cpf] else ""
            func_nome = str(row[idx_func]).strip() if idx_func is not None and row[idx_func] else "Atendente"
            forma_nome = str(row[idx_forma]).strip() if idx_forma is not None and row[idx_forma] else "Telefone"
            assunto_nome = str(row[idx_assunto]).strip() if idx_assunto is not None and row[idx_assunto] else "Geral"
            data_contato = parse_data(row[idx_data]) if idx_data is not None else date.today()
            obs = str(row[idx_obs]).strip() if idx_obs is not None and row[idx_obs] else ""

            cliente = Cliente.objects.filter(cpf=cpf).first()
            if not cliente:
                # Quarentena de contato órfão
                linha_dict = {
                    f"col_{i}": str(v) if v is not None else ""
                    for i, v in enumerate(row)
                }
                ContatoQuarentena.objects.create(
                    linha_origem=linha_dict,
                    motivo="CLIENTE_INEXISTENTE",
                    detalhe=f"CPF {cpf} não cadastrado na base de clientes.",
                )
                self.relatorio["contatos"]["quarentenados"] += 1
                continue

            func_obj, _ = Funcionario.objects.get_or_create(nome=func_nome)
            forma_obj, _ = FormaContato.objects.get_or_create(nome=forma_nome)
            assunto_obj, _ = Assunto.objects.get_or_create(nome=assunto_nome)

            Contato.objects.get_or_create(
                cliente=cliente,
                funcionario=func_obj,
                forma=forma_obj,
                assunto=assunto_obj,
                data_contato=data_contato or date.today(),
                defaults={"observacao": obs}
            )
            self.relatorio["contatos"]["criados"] += 1
