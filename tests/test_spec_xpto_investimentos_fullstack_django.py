# Testes de spec da feature xpto-investimentos-fullstack-django — ancorados via onp-spec
import os
import pytest
from datetime import date
from decimal import Decimal
from django.conf import settings
from django.core.management import call_command
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.contrib.admin.sites import site
from django.contrib.auth.models import User, Group
from django.test import RequestFactory, Client
from django.urls import reverse
from auditlog.models import LogEntry

from apps.clientes.models import Cliente, Telefone, Email, ContaBancaria, SaldoHistorico
from apps.clientes.validators import validar_cpf, formatar_cpf
from apps.clientes.admin import ClienteAdmin
from apps.clientes.permissions import mascarar_cpf_para_usuario
from apps.investimentos.models import TipoInvestimento, Investimento
from apps.relacionamento.models import Funcionario, FormaContato, Assunto, Contato, ContatoQuarentena
from apps.importacao.services import ImportadorPlanilhasService
from apps.relatorios import queries
import openpyxl


# US-001 — Fundação e Ambiente de Desenvolvimento
@pytest.mark.django_db
def test_ac_001():
    """AC-001: Ambiente de desenvolvimento executável via Docker @spec:AC-001"""
    # Dado: um ambiente configurado com settings Django
    # Quando: a configuração é verificada
    # Então: as variáveis base, apps e configurações de banco estão válidas
    installed = " ".join(settings.INSTALLED_APPS)
    assert "clientes" in installed
    assert "investimentos" in installed
    assert "relacionamento" in installed
    assert "importacao" in installed
    assert "relatorios" in installed
    assert settings.TIME_ZONE == "America/Sao_Paulo"
    assert settings.LANGUAGE_CODE == "pt-br"



# US-002 — Modelo de Dados Normalizado em 3FN e Integridade Relacional
@pytest.mark.django_db
def test_ac_002():
    """AC-002: Unicidade de CPF e dados de contato atômicos @spec:AC-002"""
    # Dado: um cliente já cadastrado com determinado CPF
    c1 = Cliente.objects.create(nome="Cliente Um", cpf="529.982.247-25")
    Telefone.objects.create(cliente=c1, numero="(11) 98765-4321", tipo="CELULAR")
    Email.objects.create(cliente=c1, endereco="cliente1@xpto.com", principal=True)

    # Quando: houver tentativa de inserção de outro cliente com o mesmo CPF
    # Então: a operação é rejeitada com aviso de duplicidade
    with pytest.raises((IntegrityError, ValidationError)):
        with transaction.atomic():
            Cliente.objects.create(nome="Cliente Dois", cpf="529.982.247-25")
    
    assert c1.telefones.count() == 1
    assert c1.emails.count() == 1


# US-002 — Modelo de Dados Normalizado em 3FN e Integridade Relacional
@pytest.mark.django_db
def test_ac_003():
    """AC-003: Unicidade temporal de saldo por conta bancária @spec:AC-003"""
    cliente = Cliente.objects.create(nome="Cliente Saldo", cpf="529.982.247-25")
    conta = ContaBancaria.objects.create(cliente=cliente, banco="XPTO", agencia="0001", conta="55555-5")
    
    # Dado: uma conta bancária com registro de saldo histórico em determinada data
    SaldoHistorico.objects.create(conta=conta, data_saldo=date(2026, 8, 1), saldo=Decimal("15000.00"))
    
    # Quando: houver tentativa de registrar um novo saldo para a mesma conta na mesma data
    # Então: a gravação é rejeitada por violação de unicidade composta (conta + data)
    with pytest.raises(IntegrityError):
        with transaction.atomic():
            SaldoHistorico.objects.create(conta=conta, data_saldo=date(2026, 8, 1), saldo=Decimal("25000.00"))


# US-002 — Modelo de Dados Normalizado em 3FN e Integridade Relacional
@pytest.mark.django_db
def test_ac_004():
    """AC-004: Proteção referencial em cascata e bloqueio de exclusão @spec:AC-004"""
    cliente = Cliente.objects.create(nome="Cliente Protegido", cpf="529.982.247-25")
    tipo = TipoInvestimento.objects.create(nome="Renda Fixa")
    Investimento.objects.create(cliente=cliente, tipo=tipo, valor_investido=Decimal("10000.00"))

    # Quando: for solicitada a exclusão de cliente com investimento vinculado
    # Então: a exclusão é bloqueada
    with pytest.raises(IntegrityError):
        with transaction.atomic():
            cliente.delete()


# US-002 — Modelo de Dados Normalizado em 3FN e Integridade Relacional
@pytest.mark.django_db
def test_ac_005():
    """AC-005: Validação de formato e dígitos verificadores de CPF @spec:AC-005"""
    # Dado: entrada com CPF inválido
    with pytest.raises(ValidationError):
        validar_cpf("111.222.333-44")  # Dígito verificador inválido
    
    # Dado: CPF válido
    validar_cpf("529.982.247-25")  # Válido


# US-003 — Interface Administrativa e Gestão Operacional
@pytest.mark.django_db
def test_ac_006():
    """AC-006: Cadastro unificado de cliente com blocos inline @spec:AC-006"""
    admin = ClienteAdmin(Cliente, site)
    inlines_classes = [inline.model for inline in admin.inlines]
    
    assert Telefone in inlines_classes
    assert Email in inlines_classes
    assert ContaBancaria in inlines_classes


# US-003 — Interface Administrativa e Gestão Operacional
@pytest.mark.django_db
def test_ac_007():
    """AC-007: Visualização de indicadores consolidados e busca rápida @spec:AC-007"""
    cliente = Cliente.objects.create(nome="Cliente Indicadores", cpf="529.982.247-25")
    conta = ContaBancaria.objects.create(cliente=cliente, banco="XPTO", agencia="0001", conta="11111-1")
    SaldoHistorico.objects.create(conta=conta, data_saldo=date(2026, 8, 1), saldo=Decimal("5000.00"))
    
    tipo = TipoInvestimento.objects.create(nome="Ações")
    Investimento.objects.create(cliente=cliente, tipo=tipo, valor_investido=Decimal("20000.00"))

    admin = ClienteAdmin(Cliente, site)
    assert cliente.saldo_atual_consolidado() == Decimal("5000.00")
    assert cliente.total_investido() == Decimal("20000.00")
    assert "nome" in admin.search_fields
    assert "cpf" in admin.search_fields


# US-004 — Controle de Acesso, Papéis e Trilha de Auditoria
@pytest.mark.django_db
def test_ac_008():
    """AC-008: Restrições de permissão e mascaramento de dados para consultores @spec:AC-008"""
    user_consultor = User.objects.create_user(username="consultor_ac8", password="p1")
    grupo_consultor, _ = Group.objects.get_or_create(name="Consultor")
    user_consultor.groups.add(grupo_consultor)

    cliente = Cliente.objects.create(nome="Cliente Mascarado", cpf="529.982.247-25")
    admin = ClienteAdmin(Cliente, site)

    request = RequestFactory().get("/admin/clientes/cliente/")
    request.user = user_consultor
    admin._current_request = request

    # Mascaramento
    cpf_exibido = admin.exibir_cpf(cliente)
    assert cpf_exibido == "529.***.***-25"

    # Bloqueio de exclusão
    assert admin.has_delete_permission(request, cliente) is False


# US-004 — Controle de Acesso, Papéis e Trilha de Auditoria
@pytest.mark.django_db
def test_ac_009():
    """AC-009: Trilha de auditoria automática em dados financeiros @spec:AC-009"""
    cliente = Cliente.objects.create(nome="Cliente Auditoria", cpf="529.982.247-25")
    tipo = TipoInvestimento.objects.create(nome="CDB Auditoria")
    inv = Investimento.objects.create(cliente=cliente, tipo=tipo, valor_investido=Decimal("10000.00"))

    # Verifica se LogEntry do django-auditlog gravou a criação
    entries = LogEntry.objects.filter(object_id=str(inv.id))
    assert entries.exists()


# US-005 — Importação e Migração Idempotente das Planilhas Legadas
@pytest.mark.django_db
def test_ac_010(tmp_path):
    """AC-010: Deduplicação e quebra de campos multivalorados na carga @spec:AC-010"""
    arquivo = tmp_path / "planilha_ac10.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["Nome", "CPF", "Telefone", "Email", "Banco", "Agência", "Conta", "Saldo", "Data"])
    ws.append(["Maria Clara", "529.982.247-25", "(11) 91111-1111 / (11) 92222-2222", "m1@x.com, m2@x.com", "XPTO", "0001", "1001-1", "10000.00", "01/08/2026"])
    ws.append(["Maria Clara", "529.982.247-25", "(11) 91111-1111", "m1@x.com", "Itaú", "0002", "2002-2", "20000.00", "01/08/2026"])
    wb.save(arquivo)

    service = ImportadorPlanilhasService(str(arquivo))
    service.executar(dry_run=False)

    assert Cliente.objects.filter(cpf="529.982.247-25").count() == 1
    cliente = Cliente.objects.get(cpf="529.982.247-25")
    assert cliente.contas.count() == 2
    assert cliente.telefones.count() == 2
    assert cliente.emails.count() == 2


# US-005 — Importação e Migração Idempotente das Planilhas Legadas
@pytest.mark.django_db
def test_ac_011(tmp_path):
    """AC-011: Quarentena de contatos órfãos e dados divergentes @spec:AC-011"""
    arquivo = tmp_path / "planilha_ac11.xlsx"
    wb = openpyxl.Workbook()
    ws_cli = wb.active
    ws_cli.title = "Clientes"
    ws_cli.append(["Nome", "CPF", "Telefone", "Email", "Banco", "Agência", "Conta", "Saldo", "Data"])
    ws_cli.append(["Joao Silva", "529.982.247-25", "(11) 91111-1111", "joao@x.com", "XPTO", "0001", "1001-1", "10000.00", "01/08/2026"])

    ws_ct = wb.create_sheet(title="Contatos")
    ws_ct.append(["CPF", "Funcionário", "Forma", "Assunto", "Data", "Observação"])
    ws_ct.append(["999.888.777-66", "Atendente", "Telefone", "Dúvida", "01/08/2026", "Contato Órfão"])
    wb.save(arquivo)

    service = ImportadorPlanilhasService(str(arquivo))
    rel = service.executar(dry_run=False)

    assert rel["contatos"]["quarentenados"] == 1
    assert ContatoQuarentena.objects.filter(motivo="CLIENTE_INEXISTENTE").count() == 1


# US-005 — Importação e Migração Idempotente das Planilhas Legadas
@pytest.mark.django_db
def test_ac_012(tmp_path):
    """AC-012: Modo de simulação sem alteração de banco @spec:AC-012"""
    arquivo = tmp_path / "planilha_ac12.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(["Nome", "CPF", "Telefone", "Email", "Banco", "Agência", "Conta", "Saldo", "Data"])
    ws.append(["Simulacao Teste", "529.982.247-25", "(11) 99999-9999", "sim@x.com", "XPTO", "0001", "9999-1", "50000.00", "01/08/2026"])
    wb.save(arquivo)

    service = ImportadorPlanilhasService(str(arquivo))
    rel = service.executar(dry_run=True)

    assert rel["clientes"]["clientes_criados"] == 1
    assert Cliente.objects.count() == 0  # Nenhuma linha persistida


# US-006 — Consultas Analíticas e Relatórios Gerenciais
@pytest.mark.django_db
def test_ac_013():
    """AC-013: Geração de relatórios gerenciais consolidados via banco @spec:AC-013"""
    c = Cliente.objects.create(nome="Cliente Relatorio AC13", cpf="529.982.247-25")
    tipo = TipoInvestimento.objects.create(nome="Tesouro Direto")
    Investimento.objects.create(cliente=c, tipo=tipo, valor_investido=Decimal("80000.00"))

    res = queries.total_investido_por_tipo()
    assert res["total_geral"] >= Decimal("80000.00")
    assert any(i["tipo"] == "Tesouro Direto" for i in res["itens"])


# US-006 — Consultas Analíticas e Relatórios Gerenciais
@pytest.mark.django_db
def test_ac_014():
    """AC-014: Bloqueio de acesso a relatórios financeiros consolidados @spec:AC-014"""
    client = Client()
    user_consultor = User.objects.create_user(username="consultor_ac14", password="p1")
    grupo_consultor, _ = Group.objects.get_or_create(name="Consultor")
    user_consultor.groups.add(grupo_consultor)
    client.force_login(user_consultor)

    response = client.get(reverse("relatorios:painel"))
    assert response.status_code == 302  # Redirecionado por não ter papel Gestor/Auditor


# US-007 — Qualidade de Código, Cobertura de Testes e Seed de Demonstração
@pytest.mark.django_db
def test_ac_015():
    """AC-015: Cobertura de testes automatizados e validação de regressão @spec:AC-015"""
    # Valida que todos os modelos principais conseguem ser instanciados e salvos
    c = Cliente.objects.create(nome="Teste Regressão", cpf="529.982.247-25")
    conta = ContaBancaria.objects.create(cliente=c, banco="XPTO", agencia="0001", conta="98765-4")
    saldo = SaldoHistorico.objects.create(conta=conta, data_saldo=date.today(), saldo=Decimal("1000.00"))
    assert c.id is not None
    assert conta.id is not None
    assert saldo.id is not None


# US-007 — Qualidade de Código, Cobertura de Testes e Seed de Demonstração
@pytest.mark.django_db
def test_ac_016():
    """AC-016: Geração de dados de demonstração coerentes @spec:AC-016"""
    # Executa o seed_demo
    call_command("seed_demo")
    assert Cliente.objects.count() >= 50
    assert ContaBancaria.objects.count() >= 50
    assert Investimento.objects.count() >= 50
    assert Contato.objects.count() >= 50
